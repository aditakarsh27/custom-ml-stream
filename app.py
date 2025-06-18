import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
import seaborn as sns
import numpy as np
import os
import sys
from openai import OpenAI
from dotenv import load_dotenv
import io
import base64
import matplotlib
matplotlib.use('Agg')

# Load environment variables
load_dotenv()

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

st.set_page_config(
    page_title="Data Chat & Visualize",
    page_icon="📊",
    layout="wide"
)

# --- Session State Initialization ---
for key, default in [
    ("messages", [{"role": "assistant", "content": "👋 Welcome! Upload a CSV or Excel file to get started, or ask me a question about your data."}]),
    ("df", None),
    ("df_description", ""),
    ("code_output", ""),
    ("interpreted_output", ""),
    ("df_info", ""),
    ("df_context", {}),
    ("file_uploader_key", 0),
    ("awaiting_response", False),
    ("current_question", ""),
    ("code", ""),
    ("namespace", {}),
    ("excel_sheets", []),  # New: List of sheet names in Excel file
    ("current_sheet", None),  # New: Currently selected sheet
    ("excel_metadata", {}),  # New: Store Excel-specific metadata
    ("is_excel", False)  # New: Flag to indicate if current file is Excel
]:
    if key not in st.session_state:
        st.session_state[key] = default

def analyze_excel_structure(excel_file):
    """Analyze Excel file structure including sheets, tables, and cell formatting"""
    try:
        # Read Excel file with openpyxl to get detailed structure
        import openpyxl
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        
        metadata = {
            "sheets": {},
            "tables": [],
            "named_ranges": []
        }
        
        # Analyze each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_info = {
                "name": sheet_name,
                "dimensions": ws.dimensions,
                "merged_cells": [str(r) for r in ws.merged_cells.ranges],
                "has_header": False,
                "potential_tables": []
            }
            
            # Detect tables and header rows
            data_region = ws.calculate_dimension()
            if ":" in data_region:
                min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(data_region)
                # Check first row for headers
                header_row = []
                for col in range(min_col, max_col + 1):
                    cell = ws.cell(min_row, col)
                    if cell.font.bold or cell.fill.start_color.index != '00000000':
                        sheet_info["has_header"] = True
                    header_row.append(cell.value)
                
                # Detect potential table regions
                if sheet_info["has_header"]:
                    sheet_info["potential_tables"].append({
                        "range": data_region,
                        "headers": header_row
                    })
            
            metadata["sheets"][sheet_name] = sheet_info
        
        return metadata
    except Exception as e:
        st.warning(f"Could not perform detailed Excel analysis: {str(e)}")
        return None

def generate_excel_description(df, metadata):
    """Generate a description of the Excel file structure and content"""
    # Get basic dataframe info
    buffer = io.StringIO()
    df.info(buf=buffer)
    df_info = buffer.getvalue()
    
    # Get sample data
    df_sample = df.head(5).to_string()
    
    # Get column descriptions
    columns_desc = df.describe().to_string()
    
    # Create Excel-specific prompt
    prompt = f"""
    I have an Excel file with the following structure and data:
    
    Excel Structure:
    - Number of sheets: {len(metadata['sheets'])}
    - Sheet names: {', '.join(metadata['sheets'].keys())}
    
    Current Sheet Data:
    DataFrame Info:
    {df_info}
    
    Sample Data (first 5 rows):
    {df_sample}
    
    Summary Statistics:
    {columns_desc}
    
    Sheet Details:
    {metadata}
    
    Please provide a comprehensive analysis including:
    1. Overall Excel file structure and organization
    2. Current sheet's data characteristics and potential purpose
    3. Data types and their appropriateness
    4. Presence of headers, tables, and special formatting
    5. Any notable patterns or relationships in the data
    6. Potential data quality issues or areas needing attention
    7. Suggestions for analysis based on the data structure
    """
    
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": prompt}],
        max_tokens=2000
    )
    
    return response.choices[0].message.content

def generate_df_description(df):
    """Generate a description of the dataframe using the OpenAI API"""
    # Get basic dataframe info
    buffer = io.StringIO()
    df.info(buf=buffer)
    df_info = buffer.getvalue()
    
    # Store df_info in session state for future use
    st.session_state.df_info = df_info
    
    # Get sample data
    df_sample = df.head(5).to_string()
    
    # Get column descriptions
    columns_desc = df.describe().to_string()
    
    # Store context in session state
    st.session_state.df_context = {
        "df_info": df_info,
        "df_sample": df_sample,
        "columns_desc": columns_desc,
        "columns": df.dtypes.to_string()
    }
    
    prompt = f"""
    I have a pandas DataFrame with the following information:
    
    DataFrame Info:
    {df_info}
    
    Sample Data (first 5 rows):
    {df_sample}
    
    Summary Statistics:
    {columns_desc}
    
    Please provide a concise summary of this dataset. Include information about:
    1. The type of data and what it might represent
    2. The number of rows and columns
    3. The data types
    4. Any notable patterns or characteristics
    """
    
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": prompt}],
        max_tokens=2000
    )
    
    return response.choices[0].message.content


def get_data_cleaning_prompt(df_info, df_sample):
    return f"""
You are an expert data scientist. The user has uploaded a dataset with the following info:

DataFrame info:
{df_info}

Sample data (first 5 rows):
{df_sample}

Write Python code to automatically clean this DataFrame:
- Handle missing values, duplicates, and obviously wrong entries.
- Encode categorical variables as needed.
- Add comments explaining each step.
- The cleaned DataFrame should be assigned to a variable called 'df_clean' by the end of the code.
- Do NOT print the whole DataFrame. Instead, print a summary of changes made.

Wrap your code in triple backticks.
    """

def get_automl_prompt(df_info, df_sample):
    return f"""
You are an expert ML engineer. The user has uploaded a cleaned DataFrame with this info:

DataFrame info:
{df_info}

Sample data (first 5 rows):
{df_sample}

Write Python code to:
- Select a suitable target column for prediction, justify your choice in comments.
- If the target is numeric, do regression; if categorical, do classification.
- Do a simple train/test split, train a model, and print test set metrics.
- Use only pandas, numpy, scikit-learn.
- Assign the trained model to a variable named 'model' and predictions to 'y_pred'.
- Encode all categorical features using one-hot encoding before fitting the model. If not required, drop the columns.
- Print only the evaluation metric(s) and a short explanation.
- Use imbalanced-learn for imbalanced datasets, and also decide other such techniques if needed based on the inforamtion about the dataset above.

Wrap your code in triple backticks.
    """


def interpret_code_output(code, output, conversation_history):
    """Use the LLM to interpret the code output in a user-friendly way"""
    # Get dataframe context
    df_info = st.session_state.df_context.get("df_info", "")
    
    # Format conversation history for the prompt
    conversation_text = ""
    for msg in conversation_history:
        conversation_text += f"{msg['role'].capitalize()}: {msg['content']}\n\n"
    
    prompt = f"""
    Given the following conversation history:
    {conversation_text}
    
    The following code was executed on a pandas DataFrame, based on the LLM's response to the user's question:
    
    ```python
    {code}
    ```
    
    The DataFrame has the following information:
    {df_info}
    
    The output of the code execution was:
    ```
    {output}
    ```
    
    Please format and display the output in a clear, concise way that would be helpful for a non-technical user. Also explain any interpretations of the output. Do not include any Certainly! or similar phrases in your response, and only include the output of the code and its explanation if it is relevant to the user's question.
    """
    
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": prompt}],
        max_tokens=1000
    )
    
    return response.choices[0].message.content

def get_error_correction_prompt(original_query, code, error, df_info, columns, df_sample, output):
    return f"""
You wrote the following Python code for the user's previous request:

User request:
{original_query}

The DataFrame has the following information:
{df_info}

Column Names and Types:
{columns}

Sample Data (first 5 rows):
{df_sample}

Code:
{code}
But when executing the code, the following error occurred:
{error}
The code generated the following output:
{output}
First think and reason about the error, then correct the code so it works as intended. Provide the COMPLETE revised Python code, wrapped in triple backticks. Reference the DataFrame as 'df'.
Add comments explaining changes if necessary.
Your Python code should reference the DataFrame as 'df'. It will be present in the namespace when the code is executed, so do not write code to create it.
"""


def process_data_query(query, df, conversation_history, max_retries=7):
    """Process a data query using the OpenAI API with conversation history, with automatic error correction and retries."""
    # Use the stored dataframe context
    df = st.session_state.namespace.get("df", df)
    buffer = io.StringIO()
    df.info(buf=buffer)
    df_info = buffer.getvalue()
    
    # Get sample data
    df_sample = df.head(5).to_string()
    
    # Get column descriptions
    columns_desc = df.describe().to_string()
    columns = df.dtypes.to_string()
    
    # Add Excel-specific context if applicable
    excel_context = ""
    if st.session_state.is_excel:
        excel_context = f"""
        This data is from an Excel file with the following structure:
        - Current sheet: {st.session_state.current_sheet}
        - Available sheets: {', '.join(st.session_state.excel_sheets)}
        - Sheet metadata: {st.session_state.excel_metadata.get('sheets', {}).get(st.session_state.current_sheet, {})}
        
        You can:
        1. Analyze relationships between different sheets
        2. Identify table structures and headers
        3. Handle merged cells and formatting
        4. Switch between sheets if needed using pandas
        5. Perform cross-sheet analysis
        """
    
    # Format conversation history for the prompt
    messages = []
    messages.append({
        "role": "system", 
        "content": f"""You are a data analysis assistant. Help the user analyze their {'Excel' if st.session_state.is_excel else 'CSV'} data.
        The DataFrame has the following information:
        {df_info}
        Column Names and Types:
        {columns}
        Sample Data (first 5 rows):
        {df_sample}
        
        {excel_context}
        
        If the user's question requires generating a visualization or performing an analysis, provide Python code using pandas, matplotlib, seaborn, or plotly.
        For visualizations, use plt.figure(figsize=(10, 6)) for matplotlib plots to ensure they're readable.
        For any results that need to be shown, print the results.
        Wrap your code in ```python and ``` tags.
        If the question is general or asking for information, provide a helpful response without code.
        Your Python code should reference the DataFrame as 'df'. It will be present in the namespace when the code is executed, so do not write code to create it.
        You have access to the complete chat history, including user queries, LLM responses, and successfully executed code.
        So, you can use variables from previous messages if needed.
        The variables present in the namespace are: {', '.join(st.session_state.namespace.keys())}
        """
    })
    
    for msg in conversation_history:
        messages.append(msg)
    messages.append({"role": "user", "content": query})

    # Main retry loop
    retries = 0
    original_query = query
    last_code = None
    last_error = None
    last_explanation = ""
    while retries <= max_retries:
        # Get response (original or correction)
        if retries == 0:
            prompt_messages = messages
        else:
            # Correction step
            df = st.session_state.namespace.get("df", df)
            buffer = io.StringIO()
            df.info(buf=buffer)
            df_info = buffer.getvalue()
            
            # Get sample data
            df_sample = df.head(5).to_string()
            
            # Get column descriptions
            columns_desc = df.describe().to_string()
            columns = df.dtypes.to_string()
            correction_prompt = get_error_correction_prompt(
                original_query, last_code, last_error, df_info, columns, df_sample, output
            )
            prompt_messages = [{
                "role": "system", "content": messages[0]["content"]
            }, {
                "role": "function", "content": correction_prompt, "name": "error_correction_prompt"
            }]
        
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=prompt_messages,
            max_tokens=2000
        )
        response_text = response.choices[0].message.content

        if "```python" in response_text:
            code_start = response_text.find("```python") + 9
            code_end = response_text.find("```", code_start)
            code = response_text[code_start:code_end].strip()
            explanation = response_text.replace("```python" + code + "```", "").strip()
            try:
                old_stdout = sys.stdout
                sys.stdout = mystdout = io.StringIO()
                local_vars = st.session_state.namespace
                if "df" not in local_vars:
                    local_vars["df"] = df
                exec(code, local_vars)
                printed_output = mystdout.getvalue()
                sys.stdout = old_stdout

                # ----------- Matplotlib -----------
                figs_matplotlib = []
                for n in plt.get_fignums():
                    fig = plt.figure(n)
                    figs_matplotlib.append(fig)

                # ----------- Plotly -----------
                fig_plotly = None
                for v in local_vars.values():
                    if str(type(v)).startswith("<class 'plotly.graph_objs._figure.Figure"):
                        fig_plotly = v
                        break
                if 'fig' in local_vars:
                    try:
                        import plotly.graph_objs
                        if isinstance(local_vars['fig'], plotly.graph_objs.Figure):
                            fig_plotly = local_vars['fig']
                    except ImportError:
                        pass

                interpreted_output = ""
                if printed_output.strip():
                    interpreted_output = interpret_code_output(code, printed_output, conversation_history)
                # SUCCESS
                st.session_state.messages.append({"role": "function", "content": response_text, "name": "final_code_explanation"})
                st.session_state.messages.append({"role": "function", "content": printed_output, "name": "final_output"})
                st.session_state.namespace = local_vars
                return {
                    "explanation": explanation,
                    "code": code,
                    "output": printed_output,
                    "interpreted_output": interpreted_output,
                    "figs_matplotlib": figs_matplotlib,
                    "fig_plotly": fig_plotly
                }
            except Exception as e:
                output = mystdout.getvalue()
                last_code = code
                last_error = str(e)
                last_explanation = explanation
                retries += 1
                continue  # Retry
        else:
            # Not code, so just return as before
            st.session_state.messages.append({"role": "assistant", "content": response_text})
            return {
                "explanation": response_text,
                "code": None,
                "output": None,
                "interpreted_output": "",
                "fig_matplotlib": None,
                "fig_plotly": None
            }

    # If we exhaust retries, return the last error
    return {
        "explanation": f"{last_explanation}\n\nError after {max_retries} retries: {last_error}",
        "code": last_code,
        "output": f"Error after {max_retries} retries: {last_error}",
        "interpreted_output": "",
        "fig_matplotlib": None,
        "fig_plotly": None
    }

def analyze_data_structure(df):
    """Analyze the first 10 rows of data to determine optimal structure and storage"""
    
    # Get the first 10 rows for analysis
    sample_data = df.head(10)
    
    # Get detailed info about the data
    buffer = io.StringIO()
    df.info(buf=buffer)
    df_info = buffer.getvalue()
    
    # Get column statistics
    column_stats = df.describe(include='all').to_string()
    
    # Create analysis prompt
    prompt = f"""
    Analyze this dataset's structure based on the first 10 rows and overall statistics:

    DataFrame Info:
    {df_info}

    First 10 rows:
    {sample_data.to_string()}

    Column Statistics:
    {column_stats}

    Please analyze:
    1. Whether the current structure is optimal
    2. If any columns should be split or merged
    3. If data types need to be changed
    4. If there are any hierarchical relationships that should be restructured
    5. If any columns contain multiple pieces of information
    6. If any numerical columns are actually categorical
    7. If date/time data needs special handling
    8. If there are any patterns in the data that suggest a different structure

    If changes are needed, provide Python code to restructure the data. The code should:
    - Create a new DataFrame with the optimal structure
    - Handle any necessary data type conversions
    - Split or merge columns as needed
    - Add comments explaining each change
    - Return the restructured DataFrame

    Wrap the code in ```python``` tags if changes are needed.
    If no changes are needed, explain why the current structure is optimal.
    """

    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": prompt}],
        max_tokens=2000
    )
    
    result = response.choices[0].message.content
    
    # If code is present in the response, execute it
    if "```python" in result:
        code_start = result.find("```python") + 9
        code_end = result.find("```", code_start)
        code = result[code_start:code_end].strip()
        
        try:
            # Create a copy of the original dataframe
            df_copy = df.copy()
            # Execute the restructuring code
            local_vars = {"df": df_copy}
            exec(code, local_vars)
            # Get the restructured dataframe
            if "df_restructured" in local_vars:
                return local_vars["df_restructured"], result
            else:
                return df_copy, result
        except Exception as e:
            st.warning(f"Error during data restructuring: {str(e)}")
            return df, result
    
    return df, result

def handle_file_upload(uploaded_file):
    try:
        file_type = uploaded_file.name.split('.')[-1].lower()
        if file_type == 'csv':
            df = pd.read_csv(uploaded_file)
            st.session_state.is_excel = False
            
            # Analyze and potentially restructure the data
            with st.spinner("Analyzing data structure..."):
                df, structure_analysis = analyze_data_structure(df)
                st.session_state.df = df
            
            # Generate description
            with st.spinner("Analyzing your data..."):
                description = generate_df_description(df)
                st.session_state.df_description = description
            
            # Add structure analysis to the message
            full_message = f"""📊 Successfully loaded **{uploaded_file.name}** with {len(df)} rows and {len(df.columns)} columns.

🔍 **Data Structure Analysis:**
{structure_analysis}

📋 **Data Description:**
{description}"""
            
            st.session_state.messages.append({
                "role": "assistant",
                "content": full_message
            })
            
        elif file_type in ['xlsx', 'xls']:
            # For Excel files, first read all sheets
            excel_file = pd.ExcelFile(uploaded_file)
            st.session_state.excel_sheets = excel_file.sheet_names
            st.session_state.is_excel = True
            
            # Read the first sheet by default
            df = pd.read_excel(excel_file, sheet_name=st.session_state.excel_sheets[0])
            st.session_state.current_sheet = st.session_state.excel_sheets[0]
            
            # Analyze Excel structure
            with st.spinner("Analyzing Excel structure..."):
                st.session_state.excel_metadata = analyze_excel_structure(uploaded_file)
            
            # Analyze and potentially restructure the data
            with st.spinner("Analyzing data structure..."):
                df, structure_analysis = analyze_data_structure(df)
                st.session_state.df = df
            
            # Generate description
            with st.spinner("Analyzing your data..."):
                description = generate_excel_description(df, st.session_state.excel_metadata)
                st.session_state.df_description = description
            
            sheet_info = ""
            if len(st.session_state.excel_sheets) > 1:
                sheet_info = f"\n\n📑 This Excel file contains {len(st.session_state.excel_sheets)} sheets: {', '.join(st.session_state.excel_sheets)}\nCurrently showing: {st.session_state.current_sheet}"
            
            # Add structure analysis to the message
            full_message = f"""📊 Successfully loaded **{uploaded_file.name}**{sheet_info}

🔍 **Data Structure Analysis:**
{structure_analysis}

📋 **Data Description:**
{description}"""
            
            st.session_state.messages.append({
                "role": "assistant",
                "content": full_message
            })
        else:
            raise ValueError(f"Unsupported file format: {file_type}")

        st.session_state.df = df

    except Exception as e:
        error_msg = f"Error reading file: {str(e)}"
        st.session_state.messages.append({"role": "assistant", "content": error_msg})

# --- Layout Start ---

st.title("Data Chat & Visualize")

# Split layout: Left = Chat, Right = Visualization
left_col, right_col = st.columns([1.7, 1.7], gap="large")

# -- LEFT COLUMN: Chat UI + File Uploader --
with left_col:
    st.header("Chat with your Data")
    chat_container = st.container()
    with chat_container:
        # Limit number of chat messages for display (e.g. last 25)
        for message in st.session_state.messages[-25:]:
            if message["role"] == "function":
                continue
            align = "assistant" if message["role"] == "assistant" else "user"
            with st.chat_message(align):
                st.markdown(message["content"])

    chat_input = st.chat_input("Ask a question about your data...")

    upload_col = st.container()
    with upload_col:
        uploaded_file = st.file_uploader(
            "Upload Data File",
            type=["csv", "xlsx", "xls"],
            key=f"file_uploader_{st.session_state.file_uploader_key}",
            help="Upload a CSV or Excel file to analyze"
        )
        
        # Excel sheet selector
        if st.session_state.is_excel and uploaded_file is not None:
            selected_sheet = st.selectbox(
                "Select Excel Sheet",
                st.session_state.excel_sheets,
                index=st.session_state.excel_sheets.index(st.session_state.current_sheet),
                key="sheet_selector"
            )
            
            if selected_sheet != st.session_state.current_sheet:
                st.session_state.current_sheet = selected_sheet
                df = pd.read_excel(uploaded_file, sheet_name=selected_sheet)
                
                # Analyze and potentially restructure the data
                with st.spinner(f"Analyzing structure of sheet: {selected_sheet}..."):
                    df, structure_analysis = analyze_data_structure(df)
                st.session_state.df = df
                
                # Generate new description for the selected sheet
                with st.spinner(f"Analyzing sheet: {selected_sheet}..."):
                    description = generate_excel_description(df, st.session_state.excel_metadata)
                    st.session_state.df_description = description
                
                # Create comprehensive message
                full_message = f"""📑 Switched to sheet: **{selected_sheet}**

🔍 **Data Structure Analysis:**
{structure_analysis}

📋 **Sheet Description:**
{description}"""
                
                st.session_state.messages.append({
                    "role": "assistant",
                    "content": full_message
                })
                st.rerun()
        
        if uploaded_file is not None:
            handle_file_upload(uploaded_file)
            st.session_state.file_uploader_key += 1
            st.rerun()

    # Handle chat input
    if chat_input:
        if st.session_state.df is None:
            st.session_state.messages.append({"role": "assistant", "content": "❗ Please upload a data file first."})
            st.rerun()
        else:
            st.session_state.awaiting_response = True
            st.session_state.current_question = chat_input
            st.session_state.messages.append({"role": "user", "content": chat_input})
            st.rerun()

# -- RIGHT COLUMN: Data Viz, Output, Tabs --
with right_col:
    st.header("Data Visualization & Analysis")
    if st.session_state.df is not None:
        viz_tabs = st.tabs(["Visualization", "Interpretation", "Code & Output"])
        # 1. Visualization
        with viz_tabs[0]:
            st.subheader("Visualization")
            # Try to display the most recent figure, if any
            figs_matplotlib = st.session_state.get("figs_matplotlib", [])
            fig_plotly = st.session_state.get("fig_plotly")
            if figs_matplotlib:
                for fig in figs_matplotlib:
                    st.pyplot(fig)
                    plt.close(fig) 
            elif fig_plotly is not None:
                st.plotly_chart(fig_plotly, use_container_width=True)
            else:
                st.info("Visualizations will appear here after you ask a question.")

        # 2. Interpretation
        with viz_tabs[1]:
            st.subheader("Interpretation")
            st.write(st.session_state.interpreted_output or "Ask a question to see an interpretation of the results.")
        # 3. Code
        with viz_tabs[2]:
            st.subheader("Code & Output")
            if st.session_state.code:
                with st.expander("View Code", expanded=True):
                    st.code(st.session_state.code, language="python")
            if st.session_state.code_output:
                with st.expander("Raw Output", expanded=True):
                    st.text(st.session_state.code_output)
            if not st.session_state.code and not st.session_state.code_output:
                st.write("Ask a question to see the code and output.")

# --- Process LLM Response ---
if st.session_state.awaiting_response and st.session_state.df is not None:
    user_query = st.session_state.current_question
    st.session_state.awaiting_response = False
    st.session_state.current_question = ""
    conversation_history = [
        {"role": msg["role"], "content": msg["content"]}
        for msg in st.session_state.messages[:-1]
        if msg["role"] in ["user", "assistant"]
    ]
    with st.spinner("Thinking..."):
        st.session_state.fig_matplotlib = None
        st.session_state.fig_plotly = None
        result = process_data_query(user_query, st.session_state.df, conversation_history)
        response_content = result["explanation"]
        if result["output"]:
            st.session_state.code_output = result["output"]
            st.session_state.interpreted_output = result["interpreted_output"]
            st.session_state.code = result["code"]
            response_content += "\n\n(See visualization and interpretation in the tabs above)"
        if result["explanation"]:
            st.write(result["explanation"])

        if result.get("code"):
            st.session_state.code = result["code"]

        # Store output for tabs
        st.session_state.code_output = result.get("output", "")
        st.session_state.interpreted_output = result.get("interpreted_output", "")
        st.session_state.figs_matplotlib = result.get("figs_matplotlib", [])
        st.session_state.fig_plotly = result.get("fig_plotly")

        st.session_state.messages.append({"role": "assistant", "content": response_content})
        st.rerun()

if __name__ == "__main__":
    pass